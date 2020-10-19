import openpyxl
from .text import camel_case, remove_punctuation, remove_parens


def get_sheet(wb, sheet_name):
    """
    Returns WorkSheet instance from provided WorkBook of sheet with given name.
    """
    sheet = wb[sheet_name]
    return sheet


def get_sheet_names(wb):
    """
    Returns list of all sheet names in provided WorkBook.
    """
    return wb.sheetnames


def get_column_letter(ws, column_number):
    """
    Provided a column number, returns its associated letter.
    """
    column_letters = [cell.column_letter for cell in ws[1]]
    column_numbers = [cell.column for cell in ws[1]]

    col_idx = column_numbers.index(column_number)
    column_letter = column_letters[col_idx]
    return column_letter


def set_cell_value(wb, sheet_name, row_num, col_num, new_value):
    """
    Set value of a particular cell.
    """
    ws = get_sheet(wb, sheet_name)
    ws.cell(row=row_num, column=col_num, value=new_value)


def new_workbook_from_sheet(wb, sheet_name):
    """
    Create new Workbook from sheet of existing workbook.
    """
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name

    ws = wb[sheet_name]
    for row in ws:
        for cell in row:
            new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
    return new_wb


def remove_rows(wb, sheet_name, start_row, end_row):
    """
    Remove one or more rows from a specified workbook's worksheet.
    """
    count = (end_row - start_row) + 1
    ws = get_sheet(wb, sheet_name)
    ws.delete_rows(start_row, count)


def remove_columns(wb, sheet_name, start_col, end_col):
    """
    Remove one or more columns from a specified workbook's worksheet.
    """
    count = (end_col - start_col) + 1
    ws = get_sheet(wb, sheet_name)
    ws.delete_cols(start_col, count)


def remove_empty_columns(wb, sheet_name):
    """
    Remove columns with no values from a WorkSheet.
    """
    ws = get_sheet(wb, sheet_name)
    keep_checking = True
    while keep_checking:
        found = False

        column_letters = [cell.column_letter for cell in ws[1]]
        column_numbers = [cell.column for cell in ws[1]]
        for idx, column_letter in enumerate(column_letters):

            col_values = [cell.value for cell in ws[column_letter]]

            all_none = all(v is None for v in col_values)
            if all_none:
                print(column_letter)
                found = True
                ws.delete_cols(column_numbers[idx], 1)
                break

        if not found:
            keep_checking = False


def remove_empty_rows(wb, sheet_name):
    """
    Remove rows in a WorkBook that have no value in any cell.
    """
    ws = get_sheet(wb, sheet_name)
    keep_checking = True
    while keep_checking:
        found = False
        for row in ws:
            row_values = [cell.value for cell in row]
            all_none = all(v is None for v in row_values)
            if all_none:

                found = True
                row_name = row[0].row
                print(row_name)
                ws.delete_rows(row_name)
                break
        if not found:
            keep_checking = False


def unmerge_row(wb, sheet_name, row_num):
    """
    Unmerge all merged cells in a row.
    Attempts to populate unmerged cells with repeated values.
    """
    ws = get_sheet(wb, sheet_name)
    # WARNING: This seems to mess up if the first row has merged cells...
    column_list = [cell.column for cell in ws[1]]

    # store all the merged columns sets
    sets = []
    # store values to populate unmerged cells with
    set_values = []
    # setup initial loop
    cur_set = []
    in_set = False
    for column_number in column_list:
        cell = ws.cell(row=row_num, column=column_number)
        if type(cell) == openpyxl.cell.cell.MergedCell:
            if in_set:
                cur_set.append(column_number)
            else:
                in_set = True
                cur_set = [column_number - 1, column_number]
                prev_cell = ws.cell(row=row_num, column=column_number - 1)
                set_values.append(prev_cell.value)
        else:
            if in_set:
                sets.append(cur_set)
                in_set = False
                cur_set = []

    print(sets)

    for idx, col_set in enumerate(sets):
        # WARNING: I assume this won't work if the merged cells span multiple rows.
        ws.unmerge_cells(
            start_row=row_num,
            start_column=col_set[0],
            end_row=row_num,
            end_column=col_set[len(col_set) - 1],
        )

        # If there is no value, will this work?
        for col_num in col_set:
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = set_values[idx]


def smear_row(wb, sheet_name, row_num):
    """
    Fills in empty cells with values from the left.
    """
    ws = get_sheet(wb, sheet_name)
    column_list = [cell.column for cell in ws[1]]

    prev_value = None

    for idx, column_number in enumerate(column_list):
        cell = ws.cell(row=row_num, column=column_number)
        value = cell.value
        if idx == 0:
            prev_value = value
        elif value is not None and value.strip() != "":
            prev_value = value
        elif value is None or value.strip() == "":
            cell.value = prev_value


def camel_case_row(wb, sheet_name, row_num):
    """
    Modify values to be camel_case.
    """
    ws = get_sheet(wb, sheet_name)
    column_list = [cell.column for cell in ws[1]]

    for idx, column_number in enumerate(column_list):
        cell = ws.cell(row=row_num, column=column_number)
        cell.value = camel_case(remove_punctuation(remove_parens(cell.value)))


def combine_rows(wb, sheet_name, row_a, row_b, combine_string=":"):
    """
    Combine values of two rows in a WorkSheet to produce a new value.
    """
    ws = get_sheet(wb, sheet_name)
    column_list = [cell.column for cell in ws[1]]

    for column_number in column_list:
        cell_a = ws.cell(row=row_a, column=column_number)
        cell_b = ws.cell(row=row_b, column=column_number)
        cell_b.value = combine_string.join([cell_a.value, cell_b.value])
    # remove top row
    ws.delete_rows(row_a)


def smear_column(wb, sheet_name, col_num):
    """
    Fills in empty cells in a column with values from the top.
    """
    ws = get_sheet(wb, sheet_name)

    column_letter = get_column_letter(ws, col_num)

    prev_value = None
    for idx, cell in enumerate(ws[column_letter]):
        value = cell.value
        if idx == 0:
            prev_value = value
        elif value is not None and value.strip() != "":
            prev_value = value
        elif value is None or value.strip() == "":
            cell.value = prev_value


def unstyle(wb, sheet_name):
    """
    Remove styling from all cells in a WorkSheet.
    """
    no_fill = openpyxl.styles.PatternFill(fill_type=None)
    side = openpyxl.styles.Side(border_style=None)
    no_border = openpyxl.styles.borders.Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    font = openpyxl.styles.Font(
        name="Arial",
        size=10,
        bold=False,
        italic=False,
        vertAlign=None,
        underline="none",
        strike=False,
        color="FF000000",
    )

    number_format = "General"

    ws = get_sheet(wb, sheet_name)
    for row in ws:
        for cell in row:
            cell.fill = no_fill
            cell.border = no_border
            cell.font = font
            cell.number_format = number_format


def convert_column_to_percent(wb, sheet_name, column_number, skip=1):
    """
    Convert provided column in WorkSheet to percent value by multiplying by 100.
    """
    ws = get_sheet(wb, sheet_name)
    column_letter = get_column_letter(ws, column_number)
    for idx, cell in enumerate(ws[column_letter]):
        if idx + 1 > skip:
            if cell.value and type(cell.value) == float:
                cell.value = cell.value * 100.0


def move_column(wb, sheet_name, column_number, new_column_number):
    """
    Move one column to new position in WorkSheet.
    """
    ws = get_sheet(wb, sheet_name)
    column_letter = get_column_letter(ws, column_number)

    row_values = []

    for idx, cell in enumerate(ws[column_letter]):
        row_values.append(cell.value)

    ws.insert_cols(new_column_number + 1)

    for idx, row_value in enumerate(row_values):
        new_cell = ws.cell(idx + 1, new_column_number + 1)
        new_cell.value = row_value

    # remove old column
    ws.delete_cols(column_number)


def move_row(wb, sheet_name, row_number, new_row_number):
    """
    Move one row to new position in WorkSheet.
    """

    ws = get_sheet(wb, sheet_name)
    column_list = [cell.column for cell in ws[1]]

    cell_values = []
    for idx, column_number in enumerate(column_list):
        cell = ws.cell(row=row_number, column=column_number)
        cell_values.append(cell.value)

    ws.insert_rows(new_row_number)

    for idx, cell_value in enumerate(cell_values):
        new_cell = ws.cell(row=new_row_number, column=idx + 1)
        new_cell.value = cell_value

    del_row_num = row_number
    if row_number > new_row_number:
        del_row_num = row_number + 1
    ws.delete_rows(del_row_num)
